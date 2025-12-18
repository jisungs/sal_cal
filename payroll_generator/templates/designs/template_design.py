# template_design.py
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    openpyxl = None

import json
import os
import logging
from pathlib import Path
from .base_design import BaseDesign

logger = logging.getLogger(__name__)

class TemplateDesign(BaseDesign):
    """템플릿 기반 디자인 기본 클래스"""
    
    def __init__(self, template_filename, mapping_filename=None):
        """
        Args:
            template_filename: 템플릿 파일명 (예: 'template_sample1.xlsx')
            mapping_filename: 셀 매핑 JSON 파일명 (선택사항)
        """
        self.template_filename = template_filename
        self.mapping_filename = mapping_filename or template_filename.replace('.xlsx', '_mapping.json')
        self.cell_mapping = self._load_cell_mapping()
        super().__init__(config_path=None)  # YAML 설정 불필요
    
    def _load_cell_mapping(self):
        """셀 매핑 JSON 파일 로드"""
        paths_to_try = []
        
        # PyInstaller 환경
        try:
            from ..utils import resource_path
            paths_to_try.append(resource_path(
                f'templates/designs/configs/{self.mapping_filename}'
            ))
        except ImportError:
            try:
                from payroll_generator.utils import resource_path
                paths_to_try.append(resource_path(
                    f'templates/designs/configs/{self.mapping_filename}'
                ))
            except ImportError:
                pass
        
        # 개발 환경
        paths_to_try.append(os.path.join(
            os.path.dirname(__file__), 'configs', self.mapping_filename
        ))
        
        for path in paths_to_try:
            if os.path.exists(path):
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        mapping_data = json.load(f)
                        return mapping_data.get('cell_mapping', {})
                except Exception as e:
                    logger.warning(f"셀 매핑 파일 로드 실패 ({path}): {e}")
                    continue
        
        logger.warning(f"셀 매핑 파일을 찾을 수 없습니다: {self.mapping_filename}. 기본 매핑 사용")
        return {}
    
    def _get_template_path(self):
        """템플릿 파일 경로 찾기
        
        경로 우선순위:
        1. templates/designs 폴더 (Railway 배포 환경 우선)
        2. sample 폴더 (로컬 개발 환경)
        3. PyInstaller resource path
        
        템플릿 파일명 매핑:
        - template_sample1.xlsx -> 급여명세서_template.xlsx (sample 폴더용)
        - template_sample2.xlsx -> 임금명세서양식_template3.xlsx (sample 폴더용)
        """
        # 템플릿 파일명 매핑 (sample 폴더용)
        template_mapping = {
            'template_sample1.xlsx': '급여명세서_template.xlsx',
            'template_sample2.xlsx': '임금명세서양식_template3.xlsx'
        }
        actual_filename = template_mapping.get(self.template_filename, self.template_filename)
        
        paths_to_try = []
        current_file = Path(__file__).resolve()
        
        # 1. templates/designs 폴더 (Railway 배포 환경 우선)
        # 이 폴더의 파일은 Git에 포함되어 Railway 배포 시 자동 포함됨
        designs_path = current_file.parent / self.template_filename
        if designs_path.exists() and designs_path.is_file():
            paths_to_try.append(str(designs_path))
            logger.debug(f"[템플릿 경로] templates/designs 폴더 경로 추가: {designs_path}")
        else:
            # 파일이 없어도 경로는 추가 (로깅용)
            paths_to_try.append(str(designs_path))
            logger.debug(f"[템플릿 경로] templates/designs 폴더에 파일 없음: {designs_path}")
        
        # 2. sample 폴더 경로 (로컬 개발 환경)
        # 프로젝트 루트 찾기
        project_root = None
        for parent in current_file.parents:
            if (parent / 'sample').exists() and (parent / 'sample' / actual_filename).exists():
                project_root = parent
                break
        
        if project_root:
            sample_path = project_root / 'sample' / actual_filename
            paths_to_try.append(str(sample_path))
            logger.debug(f"[템플릿 경로] sample 폴더 경로 추가: {sample_path}")
        
        # 3. PyInstaller 환경
        try:
            from ..utils import resource_path
            pyinstaller_path = resource_path(f'templates/designs/{self.template_filename}')
            paths_to_try.append(pyinstaller_path)
            logger.debug(f"[템플릿 경로] PyInstaller 경로 추가: {pyinstaller_path}")
        except ImportError:
            try:
                from payroll_generator.utils import resource_path
                pyinstaller_path = resource_path(f'templates/designs/{self.template_filename}')
                paths_to_try.append(pyinstaller_path)
                logger.debug(f"[템플릿 경로] PyInstaller 경로 추가 (절대 경로): {pyinstaller_path}")
            except ImportError:
                logger.debug("[템플릿 경로] PyInstaller resource_path 사용 불가")
        
        # 경로 확인 및 반환
        for path in paths_to_try:
            path_obj = Path(path)
            if path_obj.exists() and path_obj.is_file():
                logger.info(f"[템플릿 경로] 템플릿 파일 찾음: {path}")
                return str(path_obj.resolve())
        
        # 모든 경로 시도 실패
        error_msg = (
            f"템플릿 파일을 찾을 수 없습니다: {self.template_filename}.\n"
            f"다음 경로를 시도했습니다:\n"
        )
        for i, path in enumerate(paths_to_try, 1):
            exists = "✓" if Path(path).exists() else "✗"
            error_msg += f"  {i}. [{exists}] {path}\n"
        error_msg += (
            f"\n템플릿 파일이 다음 위치에 있는지 확인하세요:\n"
            f"  - payroll_generator/templates/designs/{self.template_filename} (Railway 배포용)\n"
            f"  - sample/{actual_filename} (로컬 개발용)"
        )
        logger.error(error_msg)
        raise FileNotFoundError(error_msg)
    
    def generate_excel(self, payroll_data, employee_data, output_path, period):
        """템플릿 파일을 사용하여 엑셀 생성"""
        if not HAS_OPENPYXL or openpyxl is None:
            raise ImportError(
                "openpyxl 모듈이 설치되지 않았습니다. "
                "템플릿 디자인을 사용하려면 'pip install openpyxl'을 실행하세요."
            )
        
        try:
            template_path = self._get_template_path()
            wb = openpyxl.load_workbook(template_path)
            ws = wb.active
            
            # 셀 매핑에 따라 데이터 채우기
            self._fill_template_data(ws, payroll_data, employee_data, period)
            
            # 페이지 설정: 한 페이지에 맞추기
            self._configure_page_settings(ws)
            
            # 파일 저장
            try:
                from ..utils import normalize_path
            except ImportError:
                try:
                    from payroll_generator.utils import normalize_path
                except ImportError:
                    normalize_path = lambda x: x
            
            normalized_path = normalize_path(output_path)
            wb.save(normalized_path)
            wb.close()
            
            logger.info(f"템플릿 기반 엑셀 생성 완료: {normalized_path}")
        except FileNotFoundError as e:
            logger.error(f"템플릿 파일 오류: {e}")
            raise
        except Exception as e:
            logger.error(f"엑셀 생성 실패: {e}", exc_info=True)
            raise
    
    def _safe_set_cell_value(self, ws, cell_addr, value):
        """셀에 안전하게 값 설정 (병합된 셀 처리)"""
        try:
            from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
            
            # 셀 주소 파싱
            col_letter, row = coordinate_from_string(cell_addr)
            col_idx = column_index_from_string(col_letter)
            
            # 병합된 셀 범위 확인
            for merged_range in ws.merged_cells.ranges:
                if merged_range.min_row <= row <= merged_range.max_row and \
                   merged_range.min_col <= col_idx <= merged_range.max_col:
                    # 병합된 셀인 경우, 첫 번째 셀에 값 설정
                    first_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                    first_cell.value = value
                    logger.debug(f"병합된 셀 {cell_addr} -> 첫 번째 셀 {first_cell.coordinate}에 값 설정")
                    return
            
            # 병합되지 않은 셀인 경우 직접 값 설정
            cell = ws[cell_addr]
            if hasattr(cell, 'value'):
                cell.value = value
            else:
                logger.warning(f"셀 {cell_addr}에 값을 설정할 수 없습니다 (읽기 전용).")
        except ImportError:
            # openpyxl 버전에 따라 import 경로가 다를 수 있음
            try:
                from openpyxl.utils import coordinate_from_string, column_index_from_string
                col_letter, row = coordinate_from_string(cell_addr)
                col_idx = column_index_from_string(col_letter)
                
                for merged_range in ws.merged_cells.ranges:
                    if merged_range.min_row <= row <= merged_range.max_row and \
                       merged_range.min_col <= col_idx <= merged_range.max_col:
                        first_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                        first_cell.value = value
                        return
                
                ws[cell_addr].value = value
            except Exception:
                # 최후의 수단: 직접 셀에 값 설정 시도
                try:
                    ws[cell_addr].value = value
                except Exception as e:
                    logger.warning(f"셀 {cell_addr}에 값 설정 실패: {e}")
        except Exception as e:
            logger.warning(f"셀 {cell_addr}에 값 설정 실패: {e}")
    
    def _fill_template_data(self, ws, payroll_data, employee_data, period):
        """템플릿에 데이터 채우기"""
        # 셀 매핑이 없으면 경고만 출력하고 기본 매핑 사용
        if not self.cell_mapping:
            logger.warning("셀 매핑이 비어있습니다. 기본 매핑을 사용합니다.")
            # 기본 매핑은 기존 _write_payroll_from_template과 동일한 셀 위치 사용
            self._fill_with_default_mapping(ws, payroll_data, employee_data, period)
            return
        
        # 기간 (하위 호환성을 위해 period 유지)
        if 'period' in self.cell_mapping and period:
            cell_addr = self.cell_mapping['period']
            self._safe_set_cell_value(ws, cell_addr, f"지급기간: {period}")
        
        # 지급연월 (상단 헤더)
        if 'payment_period' in self.cell_mapping and period:
            cell_addr = self.cell_mapping['payment_period']
            # period 형식이 "2025-12"인 경우 "2025년 12월"로 변환
            if period and '-' in period:
                year, month = period.split('-')
                period_formatted = f"{year}년 {month}월"
            else:
                period_formatted = period
            self._safe_set_cell_value(ws, cell_addr, period_formatted)
        
        # 직원 정보
        if 'employee_name' in self.cell_mapping:
            self._safe_set_cell_value(ws, self.cell_mapping['employee_name'], employee_data.get('이름', ''))
        
        # 소속/직급 (상단 헤더)
        if 'department_position' in self.cell_mapping:
            dept = employee_data.get('소속', employee_data.get('부서', ''))
            position = employee_data.get('직급', employee_data.get('직책', ''))
            if dept and position:
                dept_pos = f"{dept} / {position}"
            elif dept:
                dept_pos = dept
            elif position:
                dept_pos = position
            else:
                dept_pos = ''
            self._safe_set_cell_value(ws, self.cell_mapping['department_position'], dept_pos)
        
        if 'resident_number' in self.cell_mapping:
            masked_rrn = self.mask_resident_number(employee_data.get('주민번호', ''))
            # "주민번호 : 123455" 형식으로 표시
            if masked_rrn:
                formatted_rrn = f"주민번호 : {masked_rrn}"
            else:
                formatted_rrn = "주민번호 : "
            self._safe_set_cell_value(ws, self.cell_mapping['resident_number'], formatted_rrn)
        
        if 'join_date' in self.cell_mapping:
            join_date = employee_data.get('입사일', '')
            if join_date:
                if hasattr(join_date, 'strftime'):
                    self._safe_set_cell_value(ws, self.cell_mapping['join_date'], join_date.strftime('%Y-%m-%d'))
                else:
                    self._safe_set_cell_value(ws, self.cell_mapping['join_date'], str(join_date))
        
        # 지급 항목 매핑 (제목과 값 함께 표시)
        # employee_data에서도 값을 가져올 수 있도록 처리
        payment_mapping = {
            'basic_salary': ('기본급', '기본급', 'payroll'),
            'meal_allowance': ('식대', '식대', 'employee'),  # employee_data에서 가져옴
            'vehicle_maintenance': ('차량유지비', '차량유지비', 'employee'),
            'position_allowance': ('직책수당', '직책수당', 'employee'),
            'service_allowance': ('근속수당', '근속수당', 'employee'),
            'overtime': ('연장근무수당', '연장수당', 'payroll'),  # 템플릿에는 "연장수당"으로 표시됨
            'oncall_allowance': ('당직수당', '당직수당', 'employee'),
            'bonus': ('상여금', '상여금', 'payroll'),
            'other': ('기타', '기타', 'employee'),
            'total_payment': ('총지급액', '총지급액', 'payroll'),
        }
        for cell_key, (data_key, label, source) in payment_mapping.items():
            if cell_key in self.cell_mapping:
                # payroll_data 또는 employee_data에서 값 가져오기
                if source == 'payroll':
                    value = payroll_data.get(data_key, 0)
                else:  # employee
                    # employee_data에서 직접 가져오거나 0으로 설정
                    value = employee_data.get(data_key, employee_data.get(f'{data_key}원', 0))
                    if isinstance(value, str):
                        try:
                            value = int(value.replace(',', '').replace('원', ''))
                        except (ValueError, AttributeError):
                            value = 0
                    value = value if value else 0
                
                cell_addr = self.cell_mapping[cell_key]
                
                # total_payment는 H23에 숫자만 표시 (병합 셀 처리 제외)
                if cell_key == 'total_payment':
                    self._safe_set_cell_value(ws, cell_addr, value)
                    continue
                
                # 병합된 셀인 경우 제목과 값 함께 표시
                # B7:C7 형식의 병합된 셀에 "기본급 : 1,000,000원" 형식으로 표시
                from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
                try:
                    col_letter, row = coordinate_from_string(cell_addr)
                    col_idx = column_index_from_string(col_letter)
                    
                    # B열과 C열이 병합되어 있는지 확인
                    is_merged = False
                    for merged_range in ws.merged_cells.ranges:
                        if merged_range.min_row == row and \
                           merged_range.min_col == 2 and merged_range.max_col == 3:  # B:C 병합
                            is_merged = True
                            # 병합된 셀의 첫 번째 셀(B열)에 제목과 값 함께 표시
                            first_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                            # 값이 있으면 "제목 : 값원" 형식, 없으면 제목만 표시
                            if value and value > 0:
                                first_cell.value = f"{label} : {value:,}원"
                            else:
                                first_cell.value = label  # 값이 없어도 제목은 유지
                            logger.debug(f"병합된 셀 {cell_addr}에 제목과 값 함께 표시: {first_cell.value}")
                            break
                    
                    if not is_merged:
                        # 병합되지 않은 경우 기존 방식 사용
                        self._safe_set_cell_value(ws, cell_addr, value)
                except Exception as e:
                    logger.warning(f"지급 항목 {cell_key} 처리 중 오류: {e}")
                    # 폴백: 기존 방식 사용
                    self._safe_set_cell_value(ws, cell_addr, value)
        
        # 공제 항목 매핑 (제목과 값 함께 표시)
        deduction_mapping = {
            'national_pension': ('국민연금', '국민연금'),
            'health_insurance': ('건강보험', '건강보험'),
            'long_term_care': ('장기요양', '노인장기요양보험'),  # 템플릿에는 "노인장기요양보험"으로 표시됨
            'employment_insurance': ('고용보험', '고용보험'),
            'income_tax': ('소득세', '소득세'),
            'local_income_tax': ('지방소득세', '지방소득세'),
            'mutual_aid': ('상조회비', '상조회비'),  # 값이 없어도 제목은 유지
            'advance_payment': ('가불금', '가불금'),  # 값이 없어도 제목은 유지
            'total_deduction': ('총공제액', '총공제액'),
        }
        for cell_key, (data_key, label) in deduction_mapping.items():
            if cell_key in self.cell_mapping:
                value = payroll_data.get(data_key, 0)
                cell_addr = self.cell_mapping[cell_key]
                
                # total_deduction는 H24에 숫자만 표시 (병합 셀 처리 제외)
                if cell_key == 'total_deduction':
                    self._safe_set_cell_value(ws, cell_addr, value)
                    continue
                
                # 병합된 셀인 경우 제목과 값 함께 표시
                # F7:G7 형식의 병합된 셀에 "국민연금 45000" 형식으로 표시
                from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
                try:
                    col_letter, row = coordinate_from_string(cell_addr)
                    col_idx = column_index_from_string(col_letter)
                    
                    # F열과 G열이 병합되어 있는지 확인
                    is_merged = False
                    for merged_range in ws.merged_cells.ranges:
                        if merged_range.min_row == row and \
                           merged_range.min_col == 6 and merged_range.max_col == 7:  # F:G 병합
                            is_merged = True
                            # 병합된 셀의 첫 번째 셀(F열)에 제목과 값 함께 표시
                            first_cell = ws.cell(merged_range.min_row, merged_range.min_col)
                            # 값이 있으면 "제목 : 값" 형식, 없으면 "제목 : 0" 형식으로 표시
                            if value and value > 0:
                                first_cell.value = f"{label} : {value:,}"
                            else:
                                first_cell.value = f"{label} : 0"  # 값이 없어도 "제목 : 0" 형식으로 표시
                            logger.debug(f"병합된 셀 {cell_addr}에 제목과 값 함께 표시: {first_cell.value}")
                            break
                    
                    if not is_merged:
                        # 병합되지 않은 경우 기존 방식 사용
                        self._safe_set_cell_value(ws, cell_addr, value)
                except Exception as e:
                    logger.warning(f"공제 항목 {cell_key} 처리 중 오류: {e}")
                    # 폴백: 기존 방식 사용
                    self._safe_set_cell_value(ws, cell_addr, value)
        
        # 실수령액 (H25에 숫자만 표시)
        if 'net_pay' in self.cell_mapping:
            net_pay = payroll_data.get('실수령액', 0)
            # H25는 수식이 있지만, 직접 값도 써서 확실하게 표시
            self._safe_set_cell_value(ws, self.cell_mapping['net_pay'], net_pay)
    
    def _configure_page_settings(self, ws):
        """페이지 설정: 한 페이지에 맞추기"""
        try:
            # 페이지 방향: 세로 (portrait)
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            
            # 페이지 크기: A4
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            
            # 여백 최소화 (인치 단위)
            ws.page_margins.top = 0.2  # 약 5mm
            ws.page_margins.bottom = 0.2  # 약 5mm
            ws.page_margins.left = 0.2  # 약 5mm
            ws.page_margins.right = 0.2  # 약 5mm
            ws.page_margins.header = 0.0
            ws.page_margins.footer = 0.0
            
            # 인쇄 영역 설정 (B1:H28)
            ws.print_area = 'B1:H28'
            
            # 페이지에 맞추기: 너비와 높이를 1페이지에 맞춤
            # fitToWidth와 fitToHeight를 사용하면 자동으로 스케일 조정됨
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
            
            # 스케일은 fitToWidth/Height가 설정되면 무시되지만, 명시적으로 설정
            # fitToWidth/Height를 사용하면 자동으로 스케일이 조정됨
            ws.page_setup.scale = None  # None으로 설정하면 fitToWidth/Height가 활성화됨
            
            logger.debug("페이지 설정 완료: 한 페이지에 맞춤 (fitToWidth=1, fitToHeight=1)")
        except Exception as e:
            logger.warning(f"페이지 설정 중 오류 발생: {e}")
            # 페이지 설정 실패해도 계속 진행
    
    def _fill_with_default_mapping(self, ws, payroll_data, employee_data, period):
        """기본 셀 매핑 사용 (기존 _write_payroll_from_template과 동일)"""
        # 주민번호 마스킹 처리
        masked_rrn = self.mask_resident_number(employee_data.get('주민번호', ''))
        
        # 입사일 포맷팅
        join_date_str = ''
        if employee_data.get('입사일'):
            join_date = employee_data.get('입사일', '')
            if hasattr(join_date, 'strftime'):
                join_date_str = join_date.strftime('%Y-%m-%d')
            else:
                join_date_str = str(join_date)
        
        # 기간 (A2)
        if period:
            period_cell = ws['A2']
            if period_cell.value and '{PERIOD}' in str(period_cell.value):
                period_cell.value = f"지급기간: {period}"
        
        # 직원 정보 (B4, B5, B6)
        ws['B4'] = employee_data.get('이름', '')
        ws['B5'] = masked_rrn
        ws['B6'] = join_date_str
        
        # 지급 항목 (B9, B10, B11, B12)
        ws['B9'] = payroll_data.get('기본급', 0)
        ws['B10'] = payroll_data.get('연장근무수당', 0)
        ws['B11'] = payroll_data.get('상여금', 0)
        ws['B12'] = payroll_data.get('총지급액', 0)
        
        # 공제 항목 (B15~B21)
        ws['B15'] = payroll_data.get('국민연금', 0)
        ws['B16'] = payroll_data.get('건강보험', 0)
        ws['B17'] = payroll_data.get('장기요양', 0)
        ws['B18'] = payroll_data.get('고용보험', 0)
        ws['B19'] = payroll_data.get('소득세', 0)
        ws['B20'] = payroll_data.get('지방소득세', 0)
        ws['B21'] = payroll_data.get('총공제액', 0)
        
        # 실수령액 (A23)
        net_pay = payroll_data.get('실수령액', 0)
        net_pay_cell = ws['A23']
        if net_pay_cell.value and '{NET_PAY}' in str(net_pay_cell.value):
            net_pay_cell.value = f"실수령액: {net_pay:,}원"
    
    def generate_pdf(self, payroll_data, employee_data, output_path, period):
        """템플릿 엑셀 파일을 생성한 후 PDF로 변환
        
        템플릿 디자인을 유지하기 위해 엑셀 파일을 먼저 생성한 후 PDF로 변환합니다.
        엑셀→PDF 변환 라이브러리가 없으면 엑셀 파일을 생성하고 경고를 출력합니다.
        """
        if not HAS_OPENPYXL or openpyxl is None:
            raise ImportError(
                "openpyxl 모듈이 설치되지 않았습니다. "
                "템플릿 디자인을 사용하려면 'pip install openpyxl'을 실행하세요."
            )
        
        import tempfile
        import os
        
        # 1. 임시 엑셀 파일 생성
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            temp_excel_path = tmp.name
        
        try:
            # 2. 템플릿 기반 엑셀 생성
            logger.info(f"템플릿 엑셀 파일 생성 중: {temp_excel_path}")
            self.generate_excel(payroll_data, employee_data, temp_excel_path, period)
            
            # 3. 엑셀을 PDF로 변환
            # 옵션 1: LibreOffice 사용 (크로스 플랫폼, 스타일 완벽 유지, 권장)
            try:
                import subprocess
                output_dir = os.path.dirname(output_path)
                
                # 출력 디렉토리가 없으면 생성
                if output_dir and not os.path.exists(output_dir):
                    os.makedirs(output_dir, exist_ok=True)
                    logger.info(f"출력 디렉토리 생성: {output_dir}")
                
                # LibreOffice 경로 찾기
                libreoffice_cmd = None
                if os.name == 'posix':  # Mac/Linux
                    # 1. which 명령어로 PATH에서 찾기 (Nix 환경 포함)
                    try:
                        import shutil
                        libreoffice_path = shutil.which('libreoffice')
                        if libreoffice_path:
                            libreoffice_cmd = libreoffice_path
                            logger.info(f"LibreOffice 경로 발견 (which): {libreoffice_cmd}")
                        else:
                            logger.warning("LibreOffice를 PATH에서 찾을 수 없습니다 (shutil.which)")
                    except Exception as e:
                        logger.debug(f"which 명령어로 LibreOffice 찾기 실패: {e}")
                    
                    # 2. which로 찾지 못한 경우 일반적인 경로들 확인
                    if not libreoffice_cmd:
                        possible_paths = [
                            '/Applications/LibreOffice.app/Contents/MacOS/soffice',
                            '/usr/bin/libreoffice',
                            '/usr/local/bin/libreoffice',
                        ]
                        for path in possible_paths:
                            if os.path.exists(path):
                                libreoffice_cmd = path
                                logger.info(f"LibreOffice 경로 발견 (명시적 경로): {libreoffice_cmd}")
                                break
                    
                    # 3. 여전히 찾지 못한 경우 'libreoffice'로 시도 (PATH에 있을 수 있음)
                    if not libreoffice_cmd:
                        libreoffice_cmd = 'libreoffice'
                        logger.info("LibreOffice 경로를 찾지 못했습니다. 'libreoffice' 명령어로 시도합니다.")
                
                if not libreoffice_cmd:
                    raise FileNotFoundError("LibreOffice를 찾을 수 없습니다. PATH에 있는지 확인하세요.")
                
                # 한글 폰트 인식을 위한 환경 변수 설정
                env = os.environ.copy()
                
                # LibreOffice 실행 전 폰트 캐시 강제 업데이트
                try:
                    fc_cache_result = subprocess.run(
                        ['fc-cache', '-fv', '--force'],
                        capture_output=True,
                        timeout=10,
                        check=False,
                        text=True
                    )
                    if fc_cache_result.returncode == 0:
                        logger.debug("폰트 캐시 업데이트 완료")
                        # 폰트 목록 확인 (디버깅용)
                        fc_list_result = subprocess.run(
                            ['fc-list', ':', 'family'],
                            capture_output=True,
                            timeout=5,
                            check=False,
                            text=True
                        )
                        if 'Nanum' in fc_list_result.stdout or '나눔' in fc_list_result.stdout:
                            logger.info("한글 폰트(Nanum) 확인됨")
                        else:
                            logger.warning("한글 폰트가 시스템에 등록되지 않았을 수 있습니다")
                    else:
                        logger.debug(f"폰트 캐시 업데이트 실패: {fc_cache_result.stderr}")
                except FileNotFoundError:
                    logger.debug("fc-cache 명령어를 찾을 수 없습니다 (fontconfig 미설치 가능성)")
                except Exception as e:
                    logger.debug(f"폰트 캐시 업데이트 중 오류 (무시): {e}")
                
                # LibreOffice가 폰트를 인식하도록 환경 변수 설정
                # SAL_USE_VCLPLUGIN: LibreOffice의 VCL 플러그인 설정
                # FONTCONFIG_FILE: fontconfig 설정 파일 경로 (선택사항)
                env['SAL_USE_VCLPLUGIN'] = 'gen'  # 일반 VCL 플러그인 사용
                
                logger.info(f"LibreOffice를 사용하여 PDF 변환 중... (명령어: {libreoffice_cmd})")
                result = subprocess.run([
                    libreoffice_cmd, '--headless', '--convert-to', 'pdf',
                    '--outdir', output_dir,
                    temp_excel_path
                ], check=False, capture_output=True, timeout=30, env=env)
                
                # 변환 결과 확인
                if result.returncode != 0:
                    logger.warning(f"LibreOffice 변환 실패 (returncode: {result.returncode})")
                    logger.debug(f"LibreOffice stdout: {result.stdout.decode() if result.stdout else 'N/A'}")
                    logger.debug(f"LibreOffice stderr: {result.stderr.decode() if result.stderr else 'N/A'}")
                    raise subprocess.CalledProcessError(result.returncode, libreoffice_cmd, result.stdout, result.stderr)
                
                # 출력 파일명 확인 및 이동
                pdf_name = os.path.basename(temp_excel_path).replace('.xlsx', '.pdf')
                pdf_path = os.path.join(output_dir, pdf_name)
                
                # 파일 존재 여부 확인
                if not os.path.exists(pdf_path):
                    logger.error(f"LibreOffice 변환 후 PDF 파일이 생성되지 않았습니다: {pdf_path}")
                    logger.debug(f"출력 디렉토리 내용: {os.listdir(output_dir) if os.path.exists(output_dir) else '디렉토리 없음'}")
                    raise FileNotFoundError(f"PDF 파일이 생성되지 않았습니다: {pdf_path}")
                
                # 파일 이동
                if pdf_path != output_path:
                    if os.path.exists(output_path):
                        os.remove(output_path)
                    os.rename(pdf_path, output_path)
                
                # 최종 파일 존재 여부 확인
                if not os.path.exists(output_path):
                    raise FileNotFoundError(f"PDF 파일 이동 후 파일이 존재하지 않습니다: {output_path}")
                
                logger.info(f"템플릿 기반 PDF 생성 완료 (LibreOffice): {output_path}")
                return output_path
            except FileNotFoundError:
                logger.warning("LibreOffice를 찾을 수 없습니다. 폴백 로직으로 전환합니다.")
            except subprocess.TimeoutExpired:
                logger.warning("LibreOffice 변환 시간 초과. 폴백 로직으로 전환합니다.")
            except subprocess.CalledProcessError as e:
                logger.warning(f"LibreOffice 변환 실패 (returncode: {e.returncode}). 폴백 로직으로 전환합니다.")
                logger.debug(f"LibreOffice stdout: {e.stdout.decode() if e.stdout else 'N/A'}")
                logger.debug(f"LibreOffice stderr: {e.stderr.decode() if e.stderr else 'N/A'}")
            except FileNotFoundError as e:
                logger.error(f"PDF 파일 생성 실패: {e}. 폴백 로직으로 전환합니다.")
            except Exception as e:
                logger.warning(f"LibreOffice 변환 중 오류: {e}. 폴백 로직으로 전환합니다.")
                logger.exception("LibreOffice 변환 예외 상세:")
            
            # 옵션 2: Windows COM 객체 사용 (Windows 전용)
            try:
                import win32com.client
                logger.info("Windows COM 객체를 사용하여 PDF 변환 중...")
                excel = win32com.client.Dispatch("Excel.Application")
                excel.Visible = False
                workbook = excel.Workbooks.Open(os.path.abspath(temp_excel_path))
                workbook.ExportAsFixedFormat(0, os.path.abspath(output_path))  # 0 = PDF
                workbook.Close(False)
                excel.Quit()
                logger.info(f"템플릿 기반 PDF 생성 완료 (COM): {output_path}")
                return output_path
            except ImportError:
                logger.debug("win32com을 사용할 수 없습니다.")
            except Exception as e:
                logger.warning(f"COM 객체 변환 실패: {e}")
            
            # 폴백: 기본 PDF 생성기로 폴백 (무한 루프 방지를 위해 design_name=None)
            # LibreOffice가 없으므로 템플릿 디자인 PDF를 생성할 수 없음
            # 대신 엑셀 파일은 템플릿 디자인으로 생성되었으므로 사용자에게 제공
            logger.warning(
                f"엑셀→PDF 자동 변환에 실패했습니다. "
                f"LibreOffice가 설치되어 있지 않거나 사용할 수 없습니다. "
                f"기본 PDF 생성 방식으로 폴백합니다. (템플릿 엑셀 파일은 생성됨: {temp_excel_path})"
            )
            
            try:
                from ...pdf_generator import PDFGenerator
            except ImportError:
                try:
                    from payroll_generator.pdf_generator import PDFGenerator
                except ImportError:
                    from pdf_generator import PDFGenerator
            
            # 출력 디렉토리 생성 확인 (폴백 시에도 필요)
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                logger.info(f"폴백 PDF 출력 디렉토리 생성: {output_dir}")
            
            pdf_gen = PDFGenerator()
            logger.info(f"기본 PDF 생성 방식으로 폴백하여 PDF 생성 중... (design_name=None, 무한 루프 방지)")
            
            # design_name=None으로 설정하여 무한 루프 방지
            # LibreOffice가 없으므로 템플릿 디자인을 다시 시도하지 않음
            # 기본 디자인으로 PDF 생성
            try:
                result = pdf_gen.generate_payslip(
                    payroll_data, employee_data, output_path, period,
                    use_template=False, design_name=None  # 무한 루프 방지
                )
                
                # 파일 생성 확인
                if not os.path.exists(output_path):
                    logger.error(f"폴백 PDF 생성 후 파일이 존재하지 않습니다: {output_path}")
                    raise FileNotFoundError(f"폴백 PDF 생성 후 파일이 존재하지 않습니다: {output_path}")
                
                logger.warning(
                    f"PDF는 기본 디자인으로 생성되었습니다. "
                    f"템플릿 디자인이 적용된 엑셀 파일은 {temp_excel_path}에 있습니다. "
                    f"LibreOffice를 설치하면 템플릿 디자인 PDF를 생성할 수 있습니다."
                )
                logger.info(f"기본 PDF 생성 완료: {output_path}")
                return result
            except Exception as e:
                logger.error(f"폴백 PDF 생성 중 오류 발생: {e}", exc_info=True)
                # 폴백도 실패하면 예외를 다시 발생시켜 상위에서 처리하도록 함
                raise
            
        finally:
            # 임시 엑셀 파일은 디버깅을 위해 유지 (선택사항)
            # 필요시 삭제: 
            # if os.path.exists(temp_excel_path):
            #     os.unlink(temp_excel_path)
            pass
