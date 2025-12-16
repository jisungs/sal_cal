# excel_handler.py
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
try:
    from .utils import normalize_path, validate_file_path
    from .logger import setup_logger
except ImportError:
    from utils import normalize_path, validate_file_path
    from logger import setup_logger

logger = setup_logger()

class ExcelHandler:
    def __init__(self):
        self.required_columns = ['이름', '주민번호', '입사일', '기본급', '부양가족수']
    
    def read_employee_data(self, file_path):
        """직원 정보 엑셀 읽기"""
        try:
            # 파일 경로 정규화 및 검증
            validated_path = validate_file_path(file_path, ['.xlsx', '.xls'])
            logger.info(f"엑셀 파일 읽기 시작: {validated_path}")
            
            df = pd.read_excel(validated_path)
            self.validate_data(df)
            logger.info(f"엑셀 파일 읽기 완료: {len(df)}행")
            return df
        except Exception as e:
            logger.exception(f"엑셀 파일 읽기 오류: {str(e)}")
            raise ValueError(f"엑셀 파일 읽기 오류: {str(e)}")
    
    def validate_data(self, df):
        """필수 컬럼 검증"""
        missing_cols = [col for col in self.required_columns if col not in df.columns]
        if missing_cols:
            error_msg = f"필수 컬럼이 없습니다: {', '.join(missing_cols)}"
            logger.error(error_msg)
            raise ValueError(error_msg)
    
    def get_preview(self, file_path, num_rows=3):
        """엑셀 파일 미리보기 (첫 N행 반환)"""
        try:
            validated_path = validate_file_path(file_path, ['.xlsx', '.xls'])
            df = pd.read_excel(validated_path, nrows=num_rows)
            return df.to_dict('records')
        except Exception as e:
            logger.exception(f"미리보기 오류: {str(e)}")
            raise
    
    def write_payroll(self, payroll_data, output_path, employee_data, period=None, use_template=True, design_name=None):
        """급여명세서 엑셀 생성
        
        Args:
            payroll_data (dict): calculator.calculate_deductions()의 반환값
            output_path (str): 출력 파일 경로
            employee_data (dict): 직원 정보 (이름, 주민번호, 입사일 등)
            period (str, optional): 급여 기간 (예: "2025-01")
            use_template (bool): True면 템플릿 사용, False면 코드 기반 생성 (기본값: True)
            design_name (str, optional): 디자인 이름 ('template_sample1', 'template_sample2', None)
                - 'template_sample1': 급여명세서 템플릿
                - 'template_sample2': 임금명세서 템플릿
                - 'design_1', 'design_2': 더 이상 지원하지 않음 (기본 디자인으로 폴백)
        """
        # 디자인 선택 시 디자인 클래스 사용
        if design_name:
            # design_1, design_2는 더 이상 지원하지 않음
            if design_name in ['design_1', 'design_2']:
                logger.warning(
                    f"[Excel 생성] '{design_name}'는 더 이상 지원하지 않습니다. "
                    f"기본 디자인으로 폴백합니다. "
                    f"템플릿 디자인(template_sample1, template_sample2)을 사용하세요."
                )
                design_name = None  # 기본 디자인으로 폴백
            
            if design_name:  # template_sample1, template_sample2만 처리
                logger.info(f"[Excel 생성] design_name 파라미터: '{design_name}'")
                try:
                    from .templates.designs.design_factory import DesignFactory
                    logger.info(f"[Excel 생성] 디자인 팩토리에서 '{design_name}' 가져오기 시도")
                    logger.info(f"[Excel 생성] 사용 가능한 디자인: {DesignFactory.list_available_designs()}")
                    
                    design = DesignFactory.get_design(design_name)
                    logger.info(f"[Excel 생성] 디자인 인스턴스: {design is not None}")
                    
                    if design:
                        logger.info(f"[Excel 생성] 디자인 '{design_name}' 사용하여 엑셀 생성 시작")
                        result = design.generate_excel(payroll_data, employee_data, output_path, period)
                        logger.info(f"[Excel 생성] 디자인 '{design_name}' 사용하여 엑셀 생성 완료")
                        return result
                    else:
                        logger.warning(f"[Excel 생성] 디자인 '{design_name}'을 찾을 수 없습니다. 기본 방식 사용")
                        logger.warning(f"[Excel 생성] 사용 가능한 디자인: {DesignFactory.list_available_designs()}")
                except Exception as e:
                    logger.error(f"[Excel 생성] 디자인 '{design_name}' 생성 실패: {e}", exc_info=True)
                    logger.warning(f"[Excel 생성] 기본 방식으로 폴백")
        
        logger.info("[Excel 생성] 기본 방식으로 엑셀 생성")
        
        # 기존 로직 (변경 없음, 하위 호환성 유지)
        try:
            from .utils import mask_resident_number, resource_path
        except ImportError:
            from utils import mask_resident_number, resource_path
        
        try:
            # 출력 디렉토리 생성
            output_dir = os.path.dirname(output_path)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
            
            # 템플릿 사용 여부에 따라 분기
            if use_template:
                return self._write_payroll_from_template(payroll_data, output_path, employee_data, period)
            else:
                return self._write_payroll_code_based(payroll_data, output_path, employee_data, period)
        except Exception as e:
            logger.exception(f"엑셀 생성 오류: {str(e)}")
            raise ValueError(f"엑셀 생성 오류: {str(e)}")
    
    def _write_payroll_from_template(self, payroll_data, output_path, employee_data, period=None):
        """템플릿 기반 급여명세서 엑셀 생성"""
        try:
            from .utils import mask_resident_number, resource_path
        except ImportError:
            from utils import mask_resident_number, resource_path
        
        # 템플릿 파일 경로 찾기
        template_paths = [
            resource_path('templates/payroll_template.xlsx'),
            os.path.join(os.path.dirname(__file__), 'templates', 'payroll_template.xlsx'),
        ]
        
        template_path = None
        for path in template_paths:
            if os.path.exists(path):
                template_path = path
                break
        
        if not template_path:
            logger.warning("템플릿 파일을 찾을 수 없어 코드 기반 생성으로 전환합니다.")
            return self._write_payroll_code_based(payroll_data, output_path, employee_data, period)
        
        # 템플릿 파일 로드
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # 주민번호 마스킹 처리
        masked_rrn = mask_resident_number(employee_data.get('주민번호', ''))
        
        # 입사일 포맷팅
        join_date_str = ''
        if employee_data.get('입사일'):
            join_date = employee_data.get('입사일', '')
            if hasattr(join_date, 'strftime'):
                join_date_str = join_date.strftime('%Y-%m-%d')
            else:
                join_date_str = str(join_date)
        
        # 템플릿 데이터 채우기 (플레이스홀더 교체)
        # 기간 (A2)
        if period:
            period_cell = ws['A2']
            if period_cell.value and '{PERIOD}' in str(period_cell.value):
                period_cell.value = f"지급기간: {period}"
        
        # 직원 정보 (B4, B5, B6)
        ws['B4'] = employee_data.get('이름', '')
        ws['B5'] = masked_rrn
        ws['B6'] = join_date_str
        
        # 지급 항목 (B9, B10, B11, B12)
        ws['B9'] = payroll_data.get('기본급', 0)
        ws['B10'] = payroll_data.get('연장근무수당', 0)
        ws['B11'] = payroll_data.get('상여금', 0)
        ws['B12'] = payroll_data.get('총지급액', 0)
        
        # 공제 항목 (B15~B21)
        ws['B15'] = payroll_data.get('국민연금', 0)
        ws['B16'] = payroll_data.get('건강보험', 0)
        ws['B17'] = payroll_data.get('장기요양', 0)
        ws['B18'] = payroll_data.get('고용보험', 0)
        ws['B19'] = payroll_data.get('소득세', 0)
        ws['B20'] = payroll_data.get('지방소득세', 0)
        ws['B21'] = payroll_data.get('총공제액', 0)
        
        # 실수령액 (A23)
        net_pay = payroll_data.get('실수령액', 0)
        net_pay_cell = ws['A23']
        if net_pay_cell.value and '{NET_PAY}' in str(net_pay_cell.value):
            net_pay_cell.value = f"실수령액: {net_pay:,}원"
        
        # 파일 저장
        normalized_path = normalize_path(output_path)
        wb.save(normalized_path)
        wb.close()
        logger.info(f"템플릿 기반 급여명세서 엑셀 생성 완료: {normalized_path}")
    
    def _write_payroll_code_based(self, payroll_data, output_path, employee_data, period=None):
        """코드 기반 급여명세서 엑셀 생성 (기존 방식)"""
        try:
            from .utils import mask_resident_number
        except ImportError:
            from utils import mask_resident_number
        
        # 출력 디렉토리 생성
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
        
        # 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "급여명세서"
        
        # 스타일 정의
        header_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        header_font = Font(name="맑은 고딕", size=11, bold=True, color="000000")
        total_fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
        total_font = Font(name="맑은 고딕", size=11, bold=True, color="000000")
        data_font = Font(name="맑은 고딕", size=10)
        number_font = Font(name="맑은 고딕", size=10)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        # 제목
        row = 1
        ws.merge_cells(f'A{row}:B{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = "급여명세서"
        title_cell.font = Font(name="맑은 고딕", size=16, bold=True)
        title_cell.alignment = center_align
        
        # 기간
        if period:
            row += 1
            ws.merge_cells(f'A{row}:B{row}')
            period_cell = ws[f'A{row}']
            period_cell.value = f"지급기간: {period}"
            period_cell.font = data_font
            period_cell.alignment = left_align
        
        # 직원 정보
        row += 2
        ws[f'A{row}'] = "성명"
        ws[f'A{row}'].font = data_font
        ws[f'A{row}'].alignment = left_align
        ws[f'B{row}'] = employee_data.get('이름', '')
        ws[f'B{row}'].font = data_font
        ws[f'B{row}'].alignment = left_align
        
        row += 1
        ws[f'A{row}'] = "주민번호"
        ws[f'A{row}'].font = data_font
        ws[f'A{row}'].alignment = left_align
        masked_rrn = mask_resident_number(employee_data.get('주민번호', ''))
        ws[f'B{row}'] = masked_rrn
        ws[f'B{row}'].font = data_font
        ws[f'B{row}'].alignment = left_align
        
        if employee_data.get('입사일'):
            row += 1
            ws[f'A{row}'] = "입사일"
            ws[f'A{row}'].font = data_font
            ws[f'A{row}'].alignment = left_align
            join_date = employee_data.get('입사일', '')
            if hasattr(join_date, 'strftime'):
                join_date_str = join_date.strftime('%Y-%m-%d')
            else:
                join_date_str = str(join_date)
            ws[f'B{row}'] = join_date_str
            ws[f'B{row}'].font = data_font
            ws[f'B{row}'].alignment = left_align
        
        # 지급 항목
        row += 2
        ws.merge_cells(f'A{row}:B{row}')
        payment_header = ws[f'A{row}']
        payment_header.value = "지급 항목"
        payment_header.font = header_font
        payment_header.fill = header_fill
        payment_header.alignment = center_align
        payment_header.border = border
        
        row += 1
        payment_items = [
            ("기본급", payroll_data.get('기본급', 0)),
            ("연장근무수당", payroll_data.get('연장근무수당', 0)),
            ("상여금", payroll_data.get('상여금', 0)),
        ]
        
        for item_name, amount in payment_items:
            if amount > 0 or item_name == "기본급":  # 기본급은 항상 표시
                ws[f'A{row}'] = item_name
                ws[f'A{row}'].font = data_font
                ws[f'A{row}'].alignment = left_align
                ws[f'A{row}'].border = border
                
                ws[f'B{row}'] = amount
                ws[f'B{row}'].font = number_font
                ws[f'B{row}'].alignment = right_align
                ws[f'B{row}'].border = border
                ws[f'B{row}'].number_format = '#,##0'
                row += 1
        
        # 총 지급액
        ws[f'A{row}'] = "총 지급액"
        ws[f'A{row}'].font = total_font
        ws[f'A{row}'].fill = total_fill
        ws[f'A{row}'].alignment = left_align
        ws[f'A{row}'].border = border
        
        ws[f'B{row}'] = payroll_data.get('총지급액', 0)
        ws[f'B{row}'].font = total_font
        ws[f'B{row}'].fill = total_fill
        ws[f'B{row}'].alignment = right_align
        ws[f'B{row}'].border = border
        ws[f'B{row}'].number_format = '#,##0'
        
        # 공제 항목
        row += 2
        ws.merge_cells(f'A{row}:B{row}')
        deduction_header = ws[f'A{row}']
        deduction_header.value = "공제 항목"
        deduction_header.font = header_font
        deduction_header.fill = header_fill
        deduction_header.alignment = center_align
        deduction_header.border = border
        
        row += 1
        deduction_items = [
            ("국민연금", payroll_data.get('국민연금', 0)),
            ("건강보험", payroll_data.get('건강보험', 0)),
            ("장기요양", payroll_data.get('장기요양', 0)),
            ("고용보험", payroll_data.get('고용보험', 0)),
            ("소득세", payroll_data.get('소득세', 0)),
            ("지방소득세", payroll_data.get('지방소득세', 0)),
        ]
        
        for item_name, amount in deduction_items:
            ws[f'A{row}'] = item_name
            ws[f'A{row}'].font = data_font
            ws[f'A{row}'].alignment = left_align
            ws[f'A{row}'].border = border
            
            ws[f'B{row}'] = amount
            ws[f'B{row}'].font = number_font
            ws[f'B{row}'].alignment = right_align
            ws[f'B{row}'].border = border
            ws[f'B{row}'].number_format = '#,##0'
            row += 1
        
        # 총 공제액
        ws[f'A{row}'] = "총 공제액"
        ws[f'A{row}'].font = total_font
        ws[f'A{row}'].fill = total_fill
        ws[f'A{row}'].alignment = left_align
        ws[f'A{row}'].border = border
        
        ws[f'B{row}'] = payroll_data.get('총공제액', 0)
        ws[f'B{row}'].font = total_font
        ws[f'B{row}'].fill = total_fill
        ws[f'B{row}'].alignment = right_align
        ws[f'B{row}'].border = border
        ws[f'B{row}'].number_format = '#,##0'
        
        # 실수령액
        row += 2
        ws.merge_cells(f'A{row}:B{row}')
        net_pay_cell = ws[f'A{row}']
        net_pay_cell.value = f"실수령액: {payroll_data.get('실수령액', 0):,}원"
        net_pay_cell.font = Font(name="맑은 고딕", size=12, bold=True, color="0000FF")
        net_pay_cell.alignment = center_align
        net_pay_cell.border = border
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        
        # 파일 저장
        normalized_path = normalize_path(output_path)
        wb.save(normalized_path)
        logger.info(f"급여명세서 엑셀 생성 완료: {normalized_path}")

