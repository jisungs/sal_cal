#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Sample 폴더의 엑셀 템플릿 파일 분석 스크립트"""

import sys
import os

# 프로젝트 루트를 경로에 추가
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl이 설치되어 있지 않습니다. 설치 중...")
    print("pip install openpyxl 명령어로 설치해주세요.")
    sys.exit(1)

def analyze_excel(file_path):
    """엑셀 파일 분석"""
    print(f'\n{"="*70}')
    print(f'📄 {os.path.basename(file_path)} 분석')
    print(f'{"="*70}')
    
    if not os.path.exists(file_path):
        print(f'❌ 파일이 존재하지 않습니다: {file_path}')
        return
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        print(f'\n📊 기본 정보')
        print(f'  시트 개수: {len(wb.sheetnames)}')
        print(f'  시트 이름: {wb.sheetnames}')
        
        for sheet_idx, sheet_name in enumerate(wb.sheetnames, 1):
            ws = wb[sheet_name]
            print(f'\n{"-"*70}')
            print(f'📋 시트 {sheet_idx}: {sheet_name}')
            print(f'{"-"*70}')
            print(f'  크기: {ws.max_row}행 x {ws.max_column}열')
            
            # 사용된 범위 확인
            if ws.max_row > 0 and ws.max_column > 0:
                print(f'\n  📝 데이터 샘플 (첫 20행, 처음 10열):')
                print(f'  {"-"*70}')
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(20, ws.max_row), 
                                                          min_col=1, max_col=min(10, ws.max_column), 
                                                          values_only=True), 1):
                    row_data = []
                    for cell in row:
                        if cell is None:
                            row_data.append('')
                        elif isinstance(cell, (int, float)):
                            row_data.append(str(cell))
                        else:
                            cell_str = str(cell)[:30]  # 긴 텍스트는 잘라서 표시
                            row_data.append(cell_str)
                    row_str = ' | '.join(f'{cell:15}' for cell in row_data)
                    print(f'  행 {row_idx:2d}: {row_str}')
            
            # 병합된 셀 확인
            merged_cells = list(ws.merged_cells.ranges)
            if merged_cells:
                print(f'\n  🔗 병합된 셀: {len(merged_cells)}개')
                for merged in merged_cells[:10]:  # 처음 10개만 표시
                    print(f'    {merged}')
            
            # 스타일 정보 샘플 확인
            print(f'\n  🎨 스타일 정보 (샘플):')
            sample_cells = ['A1', 'B1', 'A2', 'B2']
            for cell_addr in sample_cells:
                if cell_addr in ws:
                    cell = ws[cell_addr]
                    if cell.value is not None:
                        style_info = []
                        if cell.font:
                            font_info = f"폰트:{cell.font.name or '기본'}"
                            if cell.font.size:
                                font_info += f",{cell.font.size}pt"
                            if cell.font.bold:
                                font_info += ",굵게"
                            style_info.append(font_info)
                        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                            style_info.append(f"배경:{cell.fill.start_color.rgb}")
                        if cell.alignment:
                            align_info = []
                            if cell.alignment.horizontal:
                                align_info.append(f"가로:{cell.alignment.horizontal}")
                            if cell.alignment.vertical:
                                align_info.append(f"세로:{cell.alignment.vertical}")
                            if align_info:
                                style_info.append(','.join(align_info))
                        
                        value_preview = str(cell.value)[:30] if cell.value else ''
                        print(f'    {cell_addr}: "{value_preview}" | {", ".join(style_info) if style_info else "기본 스타일"}')
            
            # 수식 확인
            wb_formula = openpyxl.load_workbook(file_path, data_only=False)
            ws_formula = wb_formula[sheet_name]
            formulas = []
            for row in ws_formula.iter_rows(min_row=1, max_row=min(50, ws_formula.max_row)):
                for cell in row:
                    if cell.data_type == 'f' and cell.value:  # 수식이 있는 셀
                        formulas.append(f"{cell.coordinate}: {cell.value}")
            
            if formulas:
                print(f'\n  🔢 수식: {len(formulas)}개 발견')
                for formula in formulas[:10]:  # 처음 10개만 표시
                    print(f'    {formula}')
            
            wb_formula.close()
        
        wb.close()
        print(f'\n{"="*70}\n')
        
    except Exception as e:
        print(f'❌ 오류 발생: {e}')
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    sample_dir = os.path.join(project_root, 'sample')
    
    files = [
        os.path.join(sample_dir, '급여명세서_template.xlsx'),
        os.path.join(sample_dir, '임금명세서양식_template3.xlsx')
    ]
    
    print('🔍 Sample 폴더 엑셀 템플릿 파일 분석 시작\n')
    
    for file_path in files:
        analyze_excel(file_path)
    
    print('✅ 분석 완료')
