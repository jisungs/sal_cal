#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""템플릿 파일 검증 스크립트"""

import os
import sys
from pathlib import Path

# 프로젝트 루트를 경로에 추가
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

try:
    import openpyxl
except ImportError:
    print("❌ openpyxl이 설치되어 있지 않습니다.")
    print("   설치 명령: pip install openpyxl")
    sys.exit(1)

def verify_template(template_path):
    """템플릿 파일 검증"""
    if not os.path.exists(template_path):
        print(f"❌ 파일이 존재하지 않습니다: {template_path}")
        return False
    
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        print(f"✓ 파일 로드 성공: {os.path.basename(template_path)}")
        print(f"  시트 이름: {ws.title}")
        print(f"  최대 행: {ws.max_row}")
        print(f"  최대 열: {ws.max_column}")
        print(f"  파일 크기: {os.path.getsize(template_path):,} bytes")
        
        # 셀에 데이터가 있는지 확인
        non_empty_cells = sum(1 for row in ws.iter_rows() for cell in row if cell.value is not None)
        print(f"  비어있지 않은 셀 수: {non_empty_cells}")
        
        wb.close()
        return True
    except Exception as e:
        print(f"❌ 파일 검증 실패: {e}")
        return False

if __name__ == '__main__':
    print("=" * 70)
    print("템플릿 파일 검증")
    print("=" * 70)
    
    templates = [
        project_root / 'sample' / '급여명세서_template.xlsx',
        project_root / 'sample' / '임금명세서양식_template3.xlsx'
    ]
    
    all_ok = True
    for template in templates:
        print()
        if not verify_template(template):
            all_ok = False
    
    print()
    print("=" * 70)
    if all_ok:
        print("✓ 모든 템플릿 파일 검증 완료")
    else:
        print("❌ 일부 템플릿 파일 검증 실패")
        sys.exit(1)
