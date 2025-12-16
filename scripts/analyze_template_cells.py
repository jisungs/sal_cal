#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""템플릿 파일의 셀 구조 분석 및 셀 매핑 자동 생성 스크립트"""

import sys
import os
import json
import re

# 프로젝트 루트를 경로에 추가
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl이 설치되어 있지 않습니다.")
    print("   설치 명령: pip install openpyxl")
    sys.exit(1)

def analyze_template(template_path, output_path=None):
    """템플릿 파일 분석 및 셀 매핑 생성
    
    Args:
        template_path: 템플릿 파일 경로
        output_path: 출력 JSON 파일 경로 (선택사항)
    
    Returns:
        dict: 셀 매핑 정보
    """
    if not os.path.exists(template_path):
        print(f"❌ 파일이 존재하지 않습니다: {template_path}")
        return None
    
    print(f"\n{'='*70}")
    print(f"📄 템플릿 분석: {os.path.basename(template_path)}")
    print(f"{'='*70}")
    
    try:
        wb = openpyxl.load_workbook(template_path, data_only=True)
        ws = wb.active
        
        cell_mapping = {}
        merged_cells = []
        formula_cells = []
        
        # 키워드 매핑 테이블
        keyword_mapping = {
            # 기간 관련
            'period': ['기간', 'period', '지급기간', '급여기간'],
            # 직원 정보
            'employee_name': ['이름', 'name', '성명', '직원명'],
            'resident_number': ['주민', 'resident', '주민번호', '등록번호'],
            'join_date': ['입사', 'join', '입사일', '입사년월일'],
            # 지급 항목
            'basic_salary': ['기본급', 'basic', '기본', '본봉'],
            'overtime': ['연장', 'overtime', '연장근무', '시간외'],
            'bonus': ['상여', 'bonus', '상여금', '보너스'],
            'total_payment': ['총지급', 'total', '지급합계', '총액'],
            # 공제 항목
            'national_pension': ['국민연금', '연금', 'pension'],
            'health_insurance': ['건강보험', '건강', 'health'],
            'long_term_care': ['장기요양', '요양', 'long'],
            'employment_insurance': ['고용보험', '고용', 'employment'],
            'income_tax': ['소득세', 'income', '소득'],
            'local_income_tax': ['지방소득세', '지방세', 'local'],
            'total_deduction': ['총공제', '공제합계', 'deduction'],
            # 실수령액
            'net_pay': ['실수령', 'net', '수령액', '최종'],
        }
        
        # 모든 셀 순회하여 키워드 매칭
        print("\n🔍 셀 분석 중...")
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                
                cell_value = str(cell.value).strip()
                cell_addr = cell.coordinate
                cell_value_lower = cell_value.lower()
                
                # 키워드 기반 자동 매핑
                for key, keywords in keyword_mapping.items():
                    if key in cell_mapping:
                        continue  # 이미 매핑됨
                    
                    for keyword in keywords:
                        if keyword.lower() in cell_value_lower:
                            # 값이 있는 셀의 오른쪽 또는 아래 셀을 데이터 셀로 추정
                            # 예: "이름:" -> B셀, "기본급" -> B셀
                            data_cell = None
                            
                            # 같은 행에서 오른쪽 셀 확인
                            if cell.column < ws.max_column:
                                right_cell = ws.cell(cell.row, cell.column + 1)
                                if right_cell.value is None or isinstance(right_cell.value, (int, float)):
                                    data_cell = right_cell.coordinate
                            
                            # 오른쪽 셀이 없으면 아래 셀 확인
                            if not data_cell and cell.row < ws.max_row:
                                below_cell = ws.cell(cell.row + 1, cell.column)
                                if below_cell.value is None or isinstance(below_cell.value, (int, float)):
                                    data_cell = below_cell.coordinate
                            
                            # 데이터 셀을 찾지 못하면 현재 셀 사용
                            if not data_cell:
                                data_cell = cell_addr
                            
                            cell_mapping[key] = data_cell
                            print(f"  ✓ {key}: {data_cell} (찾은 위치: {cell_addr}, 값: {cell_value[:30]})")
                            break
                    
                    if key in cell_mapping:
                        break
        
        # 병합된 셀 확인
        print("\n🔗 병합된 셀 확인 중...")
        for merged_range in ws.merged_cells.ranges:
            merged_cells.append(str(merged_range))
            print(f"  - {merged_range}")
        
        # 수식 확인
        print("\n🔢 수식 확인 중...")
        wb_formula = openpyxl.load_workbook(template_path, data_only=False)
        ws_formula = wb_formula.active
        for row in ws_formula.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    formula_cells.append(cell.coordinate)
                    print(f"  - {cell.coordinate}: {cell.value}")
        wb_formula.close()
        
        result = {
            'cell_mapping': cell_mapping,
            'merged_cells': merged_cells,
            'formula_cells': formula_cells,
            'sheet_name': ws.title,
            'max_row': ws.max_row,
            'max_column': ws.max_column
        }
        
        # 결과 출력
        print(f"\n{'='*70}")
        print("📊 분석 결과")
        print(f"{'='*70}")
        print(f"\n셀 매핑 ({len(cell_mapping)}개):")
        for key, cell_addr in sorted(cell_mapping.items()):
            print(f"  {key:20s} -> {cell_addr}")
        
        print(f"\n병합된 셀: {len(merged_cells)}개")
        print(f"수식 셀: {len(formula_cells)}개")
        
        # JSON 파일로 저장
        if output_path:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(result, f, ensure_ascii=False, indent=2)
            print(f"\n✅ 셀 매핑 파일 저장: {output_path}")
        
        wb.close()
        return result
        
    except Exception as e:
        print(f"❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("사용법: python scripts/analyze_template_cells.py <템플릿_파일> [출력_JSON_파일]")
        print("\n예시:")
        print("  python scripts/analyze_template_cells.py sample/급여명세서_template.xlsx")
        print("  python scripts/analyze_template_cells.py sample/급여명세서_template.xlsx \\")
        print("    payroll_generator/templates/designs/configs/template_sample1_mapping.json")
        sys.exit(1)
    
    template_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    
    result = analyze_template(template_path, output_path)
    
    if result:
        print("\n✅ 분석 완료")
    else:
        print("\n❌ 분석 실패")
        sys.exit(1)
